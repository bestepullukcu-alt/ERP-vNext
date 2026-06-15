#!/usr/bin/env python3
"""verify_module_id.py — Module ID Canonicalization Gate (DCP-002).

The Blueprint workbook
    docs/System Capability & Implementation Blueprint - master 5.xlsx :: Blueprint_Data
is the canonical authority for MOD IDs and canonical capability names.

Modes
-----
  --check-id MOD-XXXX --name "Canonical Name" [--parent MOD-YYYY] [--repo-only]
      Fail-closed preflight for a single new/changed canonical ID.
      Exit 0 = proven against Blueprint/registry.  Exit 2 = BLOCKED.

  --check-all
      Repository audit. Lists the unresolved legacy/reservation backlog and
      filename/id warnings (advisory). Exit 0 normally; exit 1 only on a NEW
      hard violation (a duplicate canonical ID). The known pre-existing backlog
      is reported, never silently ignored, and never blocks the safe migration.

Authority: execution/portfolio/delivery-capability-packs/DCP-002-module-identity-canonicalization.md
"""
import argparse
import glob
import os
import re
import sys

BLUEPRINT = "docs/System Capability & Implementation Blueprint - master 5.xlsx"
REGISTRY = "execution/registries/module-id-registry.md"
PACK_GLOB = "execution/domains/*/module-packs/*.md"


def norm(s):
    s = re.sub(r"\s+", " ", str(s or "").strip().lower())
    return re.sub(r"[^a-z0-9 ]", "", s).strip()


def load_blueprint(root):
    """Return {MOD-ID: name} or None if the workbook cannot be read (fail-closed)."""
    path = os.path.join(root, BLUEPRINT)
    if not os.path.exists(path):
        return None
    try:
        import openpyxl
    except ImportError:
        return None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None
    if "Blueprint_Data" not in wb.sheetnames:
        return {}
    rows = list(wb["Blueprint_Data"].iter_rows(values_only=True))
    if not rows:
        return {}
    header = list(rows[0])
    try:
        ci = header.index("Module ID")
        cn = header.index("Module Name")
    except ValueError:
        return {}
    bp = {}
    for r in rows[1:]:
        if ci < len(r) and r[ci] and re.match(r"^MOD-\d+$", str(r[ci]).strip()):
            bp[str(r[ci]).strip()] = str(r[cn]).strip() if cn < len(r) and r[cn] else ""
    return bp


def load_registry(root):
    path = os.path.join(root, REGISTRY)
    rows = []
    if not os.path.exists(path):
        return rows
    for line in open(path, encoding="utf-8"):
        s = line.strip()
        if not s.startswith("|"):
            continue
        c = [x.strip() for x in s.strip("|").split("|")]
        if len(c) < 8 or c[0] in ("Canonical ID",) or set(c[0]) <= set("-"):
            continue
        rows.append({
            "id": c[0], "name": c[1], "type": c[3], "status": c[4],
            "alias": c[5], "replacement": c[6],
            "notes": c[8] if len(c) > 8 else "",
        })
    return rows


def is_deprecated(row):
    return "deprecated" in row["status"].lower() or "alias" in row["type"].lower() \
        or "retired" in row["status"].lower() or "standard" in row["type"].lower()


def load_packs(root):
    out = {}
    for p in glob.glob(os.path.join(root, PACK_GLOB)):
        if os.path.basename(p).lower() == "readme.md":
            continue
        fid = None
        try:
            with open(p, encoding="utf-8") as f:
                in_fm = False
                for ln in f:
                    if ln.strip() == "---":
                        if in_fm:
                            break
                        in_fm = True
                        continue
                    if in_fm:
                        m = re.match(r"\s*id:\s*(\S+)", ln)
                        if m:
                            fid = m.group(1).strip()
        except OSError:
            pass
        out[p] = fid
    return out


def split_id(mid):
    """Return (base, parent_if_child). FU/child suffix -> parent is the base."""
    m = re.match(r"^(MOD-\d+)(-FU\w+|-[A-Za-z0-9]+)$", mid)
    if m:
        return m.group(1), m.group(1)
    if re.match(r"^MOD-\d+$", mid):
        return mid, None
    return mid, None


def check_id(root, mid, name, parent, repo_only):
    bp = load_blueprint(root)
    reg = load_registry(root)
    fails = []

    active = [r for r in reg if not is_deprecated(r)]
    if [r["id"] for r in active].count(mid) > 1:
        fails.append(f"duplicate canonical id {mid} in registry")
    for r in active:
        if r["id"] == mid and norm(r["name"]) and norm(r["name"]) != norm(name):
            fails.append(f"registry collision: {mid} already = '{r['name']}' (!= '{name}')")

    base, inferred_parent = split_id(mid)
    parent = parent or inferred_parent

    if bp is None:
        fails.append("Blueprint workbook unreadable/openpyxl missing — cannot prove ID (fail-closed)")
        return _verdict(mid, fails)

    if inferred_parent:
        known = parent in bp or any(r["id"] == parent for r in reg)
        if not known:
            fails.append(f"parent {parent} not found in Blueprint or registry")

    if base in bp:
        if not inferred_parent and norm(name) and norm(bp[base]) and norm(name) != norm(bp[base]):
            fails.append(f"canonical-name mismatch: Blueprint {base} = '{bp[base]}' (!= '{name}')")
    else:
        if repo_only:
            evidence = [r for r in reg if r["id"] == mid
                        and ("reserv" in r["notes"].lower() or "reserv" in r["status"].lower())]
            if not evidence:
                fails.append(f"repo-only {mid} lacks explicit EA reservation evidence in registry")
        else:
            fails.append(f"{base} not in Blueprint and --repo-only not set — cannot prove ID")

    return _verdict(mid, fails)


def _verdict(mid, fails):
    if fails:
        print(f"BLOCKED  {mid}")
        for f in fails:
            print(f"   - {f}")
        print("Gate failed closed. See DCP-002.")
        return 2
    print(f"OK  {mid}: proven against Blueprint/registry.")
    return 0


def runtime_hits(root, token):
    import subprocess
    for tool in (["rg", "-l", "--"], ["grep", "-rl", "--"]):
        try:
            out = subprocess.run(tool + [token, "services", "frontend", "gateway", "tests"],
                                 cwd=root, capture_output=True, text=True, timeout=30)
            return [l for l in out.stdout.splitlines() if l.strip()]
        except FileNotFoundError:
            continue
        except Exception:
            return []
    return []


def check_candidate(root, cid, name):
    """Fail-closed gate for a temporary CAND-CAP governance identity (DCP-002)."""
    fails = []
    if not re.match(r"^CAND-CAP-\d{4}(-FU\d+)?$", cid):
        fails.append("candidate ID must match CAND-CAP-#### (optional -FUn)")
    if cid.startswith("MOD-"):
        fails.append("a candidate ID can never be a Blueprint MOD id")
    reg = load_registry(root)
    rows = [r for r in reg if r["id"] == cid]
    if not rows:
        fails.append(f"{cid} has no registry row")
    active = [r for r in rows if "candidate" in r["type"].lower()]
    if len(active) > 1:
        fails.append(f"duplicate candidate id {cid}")
    for r in active:
        if norm(r["name"]) and norm(r["name"]) != norm(name):
            fails.append(f"conflicting candidate name: registry {cid} = '{r['name']}'")
        if "pending-ea" not in r["status"].lower() and "candidate" not in r["status"].lower():
            fails.append(f"{cid} is not marked temporary/pending-EA")
    recon = os.path.join(root, "execution/portfolio/blueprint-master-plan-reconciliation.md")
    if not (os.path.exists(recon) and cid in open(recon, encoding="utf-8").read()):
        fails.append(f"{cid} not recorded in the reconciliation ledger")
    hits = runtime_hits(root, cid)
    if hits:
        fails.append(f"candidate {cid} must NOT appear in runtime paths: {hits[:3]}")
    if fails:
        print(f"BLOCKED  candidate {cid}")
        for f in fails:
            print(f"   - {f}")
        print("Candidate gate failed closed. See DCP-002.")
        return 2
    print(f"OK  candidate {cid}: temporary governance identity, pending EA, not Blueprint-backed, not in runtime.")
    return 0


def check_all(root):
    reg = load_registry(root)
    bp = load_blueprint(root)
    packs = load_packs(root)

    active = [r for r in reg if not is_deprecated(r)]
    seen = {}
    for r in active:
        seen[r["id"]] = seen.get(r["id"], 0) + 1
    hard = [f"duplicate canonical id: {i} ({n}x)" for i, n in seen.items() if n > 1]

    warnings = []
    for p, fid in packs.items():
        fn = os.path.basename(p)[:-3]
        if fid and fn != fid and not fn.startswith(fid + "-"):
            warnings.append(f"filename/id: {os.path.basename(p)} has frontmatter id '{fid}'")

    backlog = []
    for r in active:
        if re.match(r"^(PSS|NEW)-", r["id"]):
            backlog.append(f"{r['id']} ({r['name']}) — legacy prefix; pending EA reservation")
        elif re.match(r"^MOD-\d+", r["id"]) and bp is not None and split_id(r["id"])[0] not in bp:
            backlog.append(f"{r['id']} ({r['name']}) — repo-only MOD not in Blueprint; needs EA reservation/confirm")

    print("== verify_module_id --check-all ==")
    print(f"registry active rows: {len(active)} / {len(reg)} | packs: {len(packs)} | "
          f"blueprint: {'loaded' if bp else 'UNAVAILABLE'}")
    candidates = [f"{r['id']} ({r['name']})" for r in reg if "candidate" in r["type"].lower()]
    compat = sorted({r["id"] for r in reg if "runtime compat" in r["notes"].lower()})
    active_legacy = sorted({r["id"] for r in active if re.match(r"^(PSS|NEW)-\d", r["id"])})
    print(f"\n[active candidate (CAND-CAP) records: {len(candidates)} — temporary, pending EA MOD-xxxx]")
    for c in sorted(candidates):
        print(f"   - {c}")
    print(f"\n[runtime compatibility literals (governance-aliased, do NOT rename in code): {len(compat)}]")
    for c in compat:
        print(f"   - {c}")
    print(f"\n[invalid ACTIVE legacy PSS-*/NEW-* canonical records (must be 0): {len(active_legacy)}]")
    for c in active_legacy:
        print(f"   - {c}")
    print(f"\n[unresolved legacy/reservation backlog: {len(backlog)} — advisory, does not block]")
    for b in sorted(backlog):
        print(f"   - {b}")
    print(f"\n[filename/id warnings: {len(warnings)} — advisory]")
    for w in sorted(warnings):
        print(f"   - {w}")
    print(f"\n[HARD violations: {len(hard)}]")
    for h in hard:
        print(f"   - {h}")
    return 1 if hard else 0


def main():
    ap = argparse.ArgumentParser(description="Module ID Canonicalization Gate (DCP-002)")
    ap.add_argument("root", nargs="?", default=".")
    ap.add_argument("--check-id")
    ap.add_argument("--name", default="")
    ap.add_argument("--parent")
    ap.add_argument("--repo-only", action="store_true")
    ap.add_argument("--candidate")
    ap.add_argument("--check-all", action="store_true")
    a = ap.parse_args()
    if a.check_all:
        sys.exit(check_all(a.root))
    if a.candidate:
        if not a.name:
            print("ERROR: --name is required with --candidate (fail-closed).")
            sys.exit(2)
        sys.exit(check_candidate(a.root, a.candidate, a.name))
    if a.check_id:
        if not a.name:
            print("ERROR: --name is required with --check-id (fail-closed).")
            sys.exit(2)
        sys.exit(check_id(a.root, a.check_id, a.name, a.parent, a.repo_only))
    ap.print_help()
    sys.exit(2)


if __name__ == "__main__":
    main()
